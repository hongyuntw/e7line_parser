#!/usr/bin/env python
# coding: utf-8

# In[4]:


from bs4 import BeautifulSoup
import os
import re
from datetime import datetime, timedelta
import requests
import json
import urllib.request, json 
import ast 
import demjson
from fake_useragent import UserAgent
from selenium import webdriver
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import openpyxl

from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
import selenium.webdriver.support.ui as ui
from time import sleep
from openpyxl.styles import Alignment


import pickle
import errno
from openpyxl.utils import get_column_letter

import smtplib
from os.path import basename
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate

error_product_url = []

def save_obj(obj,path):
    if not os.path.exists(os.path.dirname(path)):
        try:
            os.makedirs(os.path.dirname(path))
        except OSError as exc: # Guard against race condition
            if exc.errno != errno.EEXIST:
                raise
    with open(path, 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)

def load_obj(path):
    with open(path, 'rb') as f:
        return pickle.load(f)




def parseMomoProduct(urls,supplier_codes,supplier_names,product_infos):
    ua = UserAgent()
    headers = {
        'User-Agent': 'Googlebot',
    }
    product_dict = {}
    product_name_dict = {}
    product_supplier_name_dict = {}
    product_info_dict = {}
    for i in range(len(urls)):
        url = urls[i]
        print(url)
        supplier_code =  supplier_codes[i]
        supplier_name = supplier_names[i]
        product_info = product_infos[i]
        product_key = url[url.rfind('=')+1:]
        res = requests.get(url,headers=headers,allow_redirects=False)
        soup  = BeautifulSoup(res.text)
        product_div = soup.find('div', attrs={"class": "prdnoteArea"})
        if product_div is None:
            error_product_url.append(url)
            continue
        try:
            product_name = product_div.find('h1').text
        except:
            product_name = product_div.find('h3').text
        
#         dict info
        product_name_dict[product_key] = product_name
        product_supplier_name_dict[product_key] = supplier_name
        product_info_dict[product_key] = product_info
        
        spec_divs = product_div.find_all('div',attrs={'name':re.compile(r'spec*')})
        spec_div_count = len(spec_divs)

        price_list = product_div.find('ul',attrs={'class':'prdPrice'}).find_all('li')

        if(len(price_list)==2):
            original_price = int(re.findall('\d+', price_list[0].text.replace(',','') )[0])
            discount_price =  int(re.findall('\d+', price_list[1].text.replace(',','') )[0])
        else:
            original_price = ''
            discount_price = int(re.findall('\d+', price_list[0].text.replace(',','') )[0])
            
        spec_dict = {}

        # 沒有型別的
        if(spec_div_count == 0):
            tmp_dict = {}
            spec_name = "None"
            tmp_dict["discount_price"] = discount_price
            tmp_dict["original_price"] = original_price
            try:
                quantity_node = product_div.find("input", attrs={'type':'hidden','id':'goodsDtCount_001'})
                tmp_dict["quantity"] = int(quantity_node.get('value').replace(',',''))
            except:
                try:
                    quantity_node = product_div.find("input", attrs={'type':'hidden','id':'goodsDtCount_000'})
                    tmp_dict["quantity"] = int(quantity_node.get('value').replace(',',''))
                except:
                    tmp_dict["quantity"] = ''
            spec_dict[spec_name] = tmp_dict
            product_dict[product_key] = spec_dict
        else:
        #     型別超過一種
            spec_dict = {}
            for spec_div in spec_divs:
        #   
                specs = spec_div.find_all('li')
                for spec in specs:
                    all_vals = spec['val'].replace(' ','').split(',')
                    for val in all_vals:
                        spec_name = spec_dict.get(val,None)
                        if(spec_name):
                            spec_dict[val] = spec_name + " " + spec.text
                        else:
                            spec_dict[val] = spec.text

        #     print(spec_dict)
            spec_arr  = []
#             here record to save into product dict
            saved_spec_dict = {}
            for (key,name) in spec_dict.items():
                tmp_dict = {}
                momo_prdid = 'goodsDtCount_' + key
                quantity_node = product_div.find("input", attrs={'type':'hidden','id':momo_prdid})
                tmp_dict["quantity"] = int(quantity_node.get('value').replace(',',''))
                tmp_dict["discount_price"] = discount_price
                tmp_dict["original_price"] = original_price
                saved_spec_dict[name] = tmp_dict
            product_dict[product_key] = saved_spec_dict
    return product_dict, product_supplier_name_dict , product_info_dict , product_name_dict



def processMomoData(data):
    base_url  = data['url']
    productDatas = data['productData']
    product_urls = []
    product_supplier_codes = []
    product_supplier_names = []
    product_infos = []
    for productData in productDatas:
        product_url = base_url.replace('productCode',productData['ProductCode']).replace('supplierCode',productData['SupplierCode'])
        product_urls.append(product_url)
        product_supplier_codes.append(productData['SupplierCode'])
        
        if(productData['SupplierName']): 
            product_supplier_names.append(productData['SupplierName'])
        else:
            product_supplier_names.append('momo')
        product_infos.append('\n'.join(productData['ProductInfo']))
    product_dict, product_supplier_name_dict , product_info_dict , product_name_dict = parseMomoProduct(product_urls,
                                                                                   product_supplier_codes,
                                                                                   product_supplier_names,
                                                                                   product_infos)
    today = datetime.now().strftime("%m_%d_%Y")
    save_obj(product_dict, './momo/' + today + '/product_dict.pkl')
    save_obj(product_supplier_name_dict, './momo/' + today + '/product_supplier_name_dict.pkl')
    save_obj(product_info_dict, './momo/' + today + '/product_info_dict.pkl')
    save_obj(product_name_dict, './momo/' + today + '/product_name_dict.pkl')


    return product_dict, product_supplier_name_dict , product_info_dict , product_name_dict
    



def dumpExcel(product_dict, product_supplier_name_dict , product_info_dict , product_name_dict):
    today = datetime.now().strftime("%m_%d_%Y")
    yesterday = datetime.strftime(datetime.now() - timedelta(1), "%m_%d_%Y")
    yesterday_file_error = False
    
    product_change = 0
    product_remove = 0
    product_add = 0
    spec_add = 0
    spec_remove = 0
    
    try:
        old_product_dict = load_obj('./momo/' + yesterday + '/product_dict.pkl')
        old_product_supplier_name_dict = load_obj('./momo/' + yesterday + '/product_supplier_name_dict.pkl')
        old_product_info_dict = load_obj('./momo/' + yesterday + '/product_info_dict.pkl')
        old_product_name_dict = load_obj('./momo/' + yesterday + '/product_name_dict.pkl')
    except:
        yesterday_file_error = True

    
    gray = openpyxl.styles.colors.Color(rgb='00E6E6E6')
    fill_gray = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=gray)
    green = openpyxl.styles.colors.Color(rgb='0084EB00')
    fill_green = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=green)
#     record row 
    sheet_info = {}
#     record need to copy cell's row
    be_copy_cell_rows = {}

    column_width = {1:70,2:15,3:40,4:10,5:10,6:10,7:50}
    red = Font(color=colors.RED)
    workbook  = Workbook()
    workbook_changed = Workbook()
    
    for product_key , specs in product_dict.items():
        product_id = product_key
        store_name = product_supplier_name_dict[product_key]
        sheet_name = store_name
        product_name = product_name_dict[product_key]
        try:
            sheet_changed = workbook_changed.get_sheet_by_name(sheet_name)
            sheet = workbook.get_sheet_by_name(sheet_name)
        except:
            sheet = workbook.create_sheet(sheet_name)
            sheet = workbook.get_sheet_by_name(sheet_name)
            
            sheet_changed = workbook_changed.create_sheet(sheet_name)
            sheet_changed = workbook_changed.get_sheet_by_name(sheet_name)
            
            sheet['A1'].value = '商品名稱'
            sheet['B1'].value = '商品編號'
            sheet['C1'].value = '規格'
            sheet['D1'].value = '標價'
            sheet['E1'].value = '售價'
            sheet['F1'].value = '數量'
            sheet['G1'].value = 'info'
            sheet_info[sheet_name] = 2
            
#             for 另外excel
            be_copy_cell_rows[sheet_name] = [1]
            
        need_to_copy = False
    
        row = sheet_info[sheet_name]
        col = 1
        

        sheet.cell(row = row,column = col).value = product_name
        product_info = product_info_dict[product_key]
        sheet.cell(row = row,column = 7).value = product_info

        
        col += 1
        sheet.cell(row = row,column = col).value = product_id
        col += 1
        row_start  = row
        original_spec_add = spec_add
        
        for spec_name , spec in specs.items():
            sheet.cell(row = row,column = col).value = spec_name
            col += 1
            sheet.cell(row = row,column = col).value = spec["original_price"]
            col += 1
            try:
                old_price = int(old_product_dict[product_key][spec_name]['discount_price'])
                price = int(spec["discount_price"])
                if(old_price != price):
                    product_change += 1
                    need_to_copy = True
                    be_copy_cell_rows[sheet_name].append(row)
                    sheet.cell(row = row,column = col).value = str(old_price) + "->" +  str(price)
                    sheet['E'+str(row)].font = red
                else:
                    sheet.cell(row = row, column = col).value = price
                    
            except:
                sheet.cell(row = row,column = col).value = spec["discount_price"]

            col += 1
            sheet.cell(row = row,column = col).value = spec["quantity"]
            
#             品項新增
            try:
                old_spec = old_product_dict[product_key][spec_name]
            except:
                spec_add += 1
                need_to_copy = True
                be_copy_cell_rows[sheet_name].append(row)
                for i in range(1,8):
                    sheet.cell(row = row,column = i).fill = fill_green
            row += 1
            col = 3

        if(original_spec_add != spec_add):
            product_add += 1
        
        if need_to_copy :
            if row_start not in be_copy_cell_rows[sheet_name]:
                be_copy_cell_rows[sheet_name].append(row_start)
            
      
        sheet_info[sheet_name] = row
        for r in range(1,row):
            for c in range(1,8):
                sheet.cell(row = r,column = c).font = sheet.cell(row = r,column = c).font.copy(size=14)
                sheet.column_dimensions[get_column_letter(c)].width = column_width[c]       
        
        
    
    
    if not yesterday_file_error:
        for product_key , specs in old_product_dict.items():
    #         昨天有，今天沒有，已下架
            try:
                spec_remove_flag = False
                buf = product_dict[product_key]
                spec_name_list = []
                spec_list = []
                for spec_name , spec in specs.items():
                    try:
                        buf = product_dict[product_key][spec_name]

                    except:
#                         print(spec_name)
#                         print(spec)
                        spec_name_list.append(spec_name)
                        spec_list.append(spec)
                        spec_remove += 1
                        continue
                    if(len(spec_list)!=0):
                        spec_remove_flag = True
                        raise
                        
            except:
                product_id = product_key
                store_name = old_product_supplier_name_dict[product_key]
                sheet_name = store_name
                product_name = old_product_name_dict[product_key]
                product_remove += 1

                try:
                    sheet = workbook.get_sheet_by_name(sheet_name)
                except:
                    sheet = workbook.create_sheet(sheet_name)
                    sheet = workbook.get_sheet_by_name(sheet_name)
                    
                    sheet_changed = workbook_changed.create_sheet(sheet_name)
                    sheet_changed = workbook_changed.get_sheet_by_name(sheet_name)
                    
                    sheet['A1'].value = '商品名稱'
                    sheet['B1'].value = '商品編號'
                    sheet['C1'].value = '規格'
                    sheet['D1'].value = '標價'
                    sheet['E1'].value = '售價'
                    sheet['F1'].value = '數量'
                    sheet['G1'].value = 'info'
                    sheet_info[sheet_name] = 2
                    
                    be_copy_cell_rows[sheet_name] = [1]


                row = sheet_info[sheet_name]
                row_start = row
                col = 1
                sheet.cell(row = row,column = col).value = product_name
                sheet.cell(row = row,column = 7).value = old_product_info_dict[product_key]
                col += 1
                sheet.cell(row = row,column = col).value = product_id
                col += 1
                
                
                
                
                
                if(spec_remove_flag):
                    
                    
                    for i in range(len(spec_name_list)):
                        spec = spec_list[i]
                        spec_name = spec_name_list[i]
                        sheet.cell(row = row,column = col).value = spec_name
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["original_price"]
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["discount_price"]
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["quantity"]
                        row += 1
                        col = 3
                else:


                
                    for spec_name , spec in specs.items():
                        sheet.cell(row = row,column = col).value = spec_name
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["original_price"]
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["discount_price"]
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["quantity"]
                        row += 1
                        col = 3
                        spec_remove += 1
                sheet_info[sheet_name] = row

                for r in range(row_start,row):
                    for c in range(1,8):
                        sheet.cell(row = r,column = c).font = sheet.cell(row = r,column = c).font.copy(size=14)
                        sheet.cell(row = r,column = c).fill = fill_gray
                        sheet.column_dimensions[get_column_letter(c)].width = column_width[c]

                for i in range(row_start,row):
                    be_copy_cell_rows[sheet_name].append(i)


                
    try:
        del workbook['Sheet']
        del workbook_changed['Sheet']
    except:
        pass
    
    
    for s_name, rows in be_copy_cell_rows.items():
        sheet_changed = workbook_changed.get_sheet_by_name(s_name)
        sheet = workbook.get_sheet_by_name(s_name)
        i = 1
        for row in rows:
            for col in range(1,8):
                sheet_changed.cell(row = i,column = col).value = sheet.cell(row = row,column = col).value
                sheet_changed.cell(row = i,column = col).font = sheet.cell(row = row,column = col).font.copy()
                sheet_changed.cell(row = i,column = col).fill = sheet.cell(row = row,column = col).fill.copy()
                sheet_changed.column_dimensions[get_column_letter(col)].width = column_width[col]

            i += 1
            
    
    
    
    workbook.save('./momo/'+ today + '/momo.xlsx')
    workbook_changed.save('./momo/'+ today + '/momo_changed.xlsx')
    fileNames = ['momo/'+ today + '/momo.xlsx',
                 'momo/'+ today + '/momo_changed.xlsx']
    
    today = datetime.now().strftime("%m_%d_%Y %H:%M:%S")
    text = today + ' 商品變更數量:' + str(product_change) + ', 商品移除數量: ' + (str(product_remove))
    text += ' ,品項移除數量:' + str(spec_remove)
    text += '\n總共有' + str(product_add) + '個商品的品項有新增，共比昨天多了'+ str(spec_add) + '個品項'
    text += '\n綠色代表新增，灰色代表移除，紅色代表變更！'
    for url in error_product_url:
        text += '\n' + url + ' 商品不存在'
    
    return  fileNames , text



def send_mail(send_from, send_to, subject, text, files=None,server="127.0.0.1"):
    assert isinstance(send_to, list)
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = COMMASPACE.join(send_to)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    msg.attach(MIMEText(text))
    
    for file in files:
        with open(file, "rb") as fil:
            part = MIMEApplication(
                fil.read(),
                Name=basename(file)
            )
            # After the file is closed
            part['Content-Disposition'] = 'attachment; filename="%s"' % basename(file)
            msg.attach(part)
    smtp = smtplib.SMTP('10.210.1.221')
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.close()






subject = 'Momo商品爬蟲'
send_to = ['ruby.lin@e7line.com',
            'xing.chen@gigabyte.com',
            'chaoyang.huang@gigabyte.com',
          'harrychiang0@gmail.com']
# send_to = ['harrychiang0@gmail.com']


test_api = 'https://www.e7line.com:8080/spiderdata3.aspx'
api = 'https://www.e7line.com/spiderdata3.aspx'
with urllib.request.urlopen(api) as url:
    datas = json.loads(url.read().decode())
    
try:
    momo_data = datas[0]
except:
    text = '今天momo商品沒有資訊！'
    send_mail('e7line@gigabyte.com', send_to , subject, text , files=None,server="127.0.0.1")
    print('cant get momo data')
    raise SystemExit("stop program")

product_dict, product_supplier_name_dict , product_info_dict , product_name_dict = processMomoData(momo_data)
fileNames , text = dumpExcel(product_dict, product_supplier_name_dict , product_info_dict , product_name_dict)

    
send_mail('e7line@gigabyte.com', send_to , subject, text , files=fileNames,server="127.0.0.1")





