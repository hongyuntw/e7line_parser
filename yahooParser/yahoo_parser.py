#!/usr/bin/env python
# coding: utf-8

# In[7]:


from bs4 import BeautifulSoup
import os
import re
from datetime import datetime, timedelta
import requests
import json
import urllib.request, json 
import ast 
import demjson
from fake_useragent import UserAgent
from selenium import webdriver
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import openpyxl

from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
import selenium.webdriver.support.ui as ui
from time import sleep
from openpyxl.styles import Alignment


import pickle
import errno
from openpyxl.utils import get_column_letter

import smtplib
from os.path import basename
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate


# In[8]:


def save_obj(obj,path):
    if not os.path.exists(os.path.dirname(path)):
        try:
            os.makedirs(os.path.dirname(path))
        except OSError as exc: # Guard against race condition
            if exc.errno != errno.EEXIST:
                raise
    with open(path, 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)

def load_obj(path):
    with open(path, 'rb') as f:
        return pickle.load(f)


# In[9]:








def getAllProducts(store_url):
    product_urls = []
    for i in range(1,100):
        base_page_url = '?apg='
        page_url = base_page_url + str(i)
        url = store_url + page_url
        print(url)
        res = requests.get(url,headers=headers,allow_redirects=False)
        soup  = BeautifulSoup(res.text,'lxml')
        products_nodes = soup.find('div',attrs={'class':'bd clearfix'}).find_all('li')
        if (len(products_nodes) == 0):
            return product_urls
        for product_node in products_nodes:
            product_url = product_node.find('a')['href']
            product_urls.append(product_url)
    return product_urls



def parseYahooProduct(urls,supplier_codes,supplier_names,product_infos):
    ua = UserAgent()
    headers = {
        'User-Agent': 'Googlebot',
    }
    product_dict = {}
    product_name_dict = {}
    product_supplier_name_dict = {}
    product_info_dict = {}
    for i in range(len(urls)):
        url = urls[i]
        print(url)
        supplier_code =  supplier_codes[i]
        supplier_name = supplier_names[i]
        product_info = product_infos[i]
        product_key = url[url.rfind('/')+1:]
        
        res = requests.get(url,headers=headers,allow_redirects=False)
        soup  = BeautifulSoup(res.text,'lxml')
        try:
            product_div = soup.find('section',attrs={"class":"wrap__3MZRM"})
            product_name = product_div.find('h1',attrs={'class':'title__3wBva'}).text
        except:
            product_name = '商品名稱錯誤'
#         print(product_name)
        product_name_dict[product_key] = product_name
        product_supplier_name_dict[product_key] = supplier_name
        product_info_dict[product_key] = product_info
        

#         try:
#             product_id = soup.find('input',attrs={'type':'hidden','name':'listingId'}).attrs.get('value')
#             product_id_dict[product_name] = str(product_id)
#         except:
#             product_id_dict[product_name] = '商品編號錯誤'
        
        
        original_price = ''
        discount_price = ''
        try:
            original_price_node = product_div.find('span',attrs={'class':'originPrice__271Nh'})
            original_price = int(re.findall('\d+', original_price_node.text.replace(',','') )[0])
        except:
            pass
        
        try:
            discount_price_node = product_div.find('em',attrs={'class':'price__2f7Jw'})
            discount_price = int(re.findall('\d+', discount_price_node.text.replace(',','') )[0])
        except:
            pass

        try:
            spec_ul_node = product_div.find('ul',attrs={'class':'specList__3TA_I'})
            spec_divs = spec_ul_node.find_all('li',attrs={'class':None})
            spec_div_count = len(spec_divs)
        except:
            continue
            
#         print(discount_price)
#         print(original_price)


        # 沒有型別的
        if(spec_div_count == 0):
            spec_dict = {}
            tmp_dict = {}
            spec_name = 'None'
            tmp_dict["discount_price"] = discount_price
            tmp_dict["original_price"] = original_price
            try:
                quantity_node = soup.find("input", attrs={'type':'number','class':'qtyInput__1dbgq'})
                tmp_dict["quantity"] = int(quantity_node.get('max').replace(',',''))
                if(tmp_dict['quantity']>5000):
                    tmp_dict['quantity'] = '5000+'
            except:
                tmp_dict["quantity"] = ''
            spec_dict[spec_name] = tmp_dict
                
            product_dict[product_key] = spec_dict
        else:
        #     型別超過一種
            spec_ids_arr = []
            for spec_div in spec_divs:
                lis = spec_div.find('div',attrs={'class':'content__3X3yq'}).find('ul').find_all('li')
                spec_ids = []
                for li in lis:
                    spec_id = li.find('input').attrs.get('id')
                    spec_ids.append(spec_id)
                spec_ids_arr.append(spec_ids)
            op = webdriver.ChromeOptions()
            op.add_argument('headless')
            driver = webdriver.Chrome(options=op)
            driver.get(url)
            try:
                wait = ui.WebDriverWait(driver,5)
                wait.until(lambda driver: driver.find_element_by_class_name("specCheckbox__LtDOH"))
            except:
                continue
            spec_arr = []
            spec_dict = {}
            for spec_id in spec_ids_arr[0]:
                tmp_dict = {}
                element = driver.find_element_by_id(spec_id)
#                 print(spec_id)
                if(element.get_attribute('disabled') is None ):
                    driver.execute_script("arguments[0].click();", element)
                    html = driver.page_source
                    soup = BeautifulSoup(html,'lxml')
                    try:
                        spec_name = soup.find('label',attrs={'for':spec_id}).text
                    except:
                        spec_name = 'None'

                    tmp_dict["discount_price"] = discount_price
                    tmp_dict["original_price"] = original_price
                    try:
                        quantity_node = soup.find("input", attrs={'type':'number','class':'qtyInput__1dbgq'})
                        tmp_dict["quantity"] = int(quantity_node.get('max').replace(',',''))
                        if(tmp_dict['quantity']>5000):
                            tmp_dict['quantity'] = '5000+'
                    except:
                        tmp_dict["quantity"] = ''
#                     print(tmp_dict)
                    spec_dict[spec_name] = tmp_dict
                else:
                    try:
                        spec_name = soup.find('label',attrs={'for':spec_id}).text
                    except:
                        spec_name = 'None' 
                    tmp_dict["discount_price"] = ''
                    tmp_dict["original_price"] = ''
                    tmp_dict['quantity']= ''
                    spec_dict[spec_name] = tmp_dict

                    
            product_dict[product_key] = spec_dict
            
        try:
            driver.close()
            driver.quit() 
        except:
            pass
    return product_dict, product_supplier_name_dict , product_info_dict , product_name_dict




def processYahooData(data):
    base_url  = data['url']
    productDatas = data['productData']
    product_urls = []
    product_supplier_codes = []
    product_supplier_names = []
    product_infos = []
    for productData in productDatas:
        product_url = base_url.replace('productCode',productData['ProductCode']).replace('supplierCode',productData['SupplierCode'])
        product_urls.append(product_url)
        product_supplier_codes.append(productData['SupplierCode'])
        
        if(productData['SupplierName']): 
            product_supplier_names.append(productData['SupplierName'])
        else:
            product_supplier_names.append('yahoo')
        product_infos.append('\n'.join(productData['ProductInfo']))
    product_dict, product_supplier_name_dict , product_info_dict , product_name_dict = parseYahooProduct(product_urls,
                                                                                   product_supplier_codes,
                                                                                   product_supplier_names,
                                                                                   product_infos)
    today = datetime.now().strftime("%m_%d_%Y")
    save_obj(product_dict, './yahoo/' + today + '/product_dict.pkl')
    save_obj(product_supplier_name_dict, './yahoo/' + today + '/product_supplier_name_dict.pkl')
    save_obj(product_info_dict, './yahoo/' + today + '/product_info_dict.pkl')
    save_obj(product_name_dict, './yahoo/' + today + '/product_name_dict.pkl')


    return product_dict, product_supplier_name_dict , product_info_dict , product_name_dict
    



def dumpExcel(product_dict, product_supplier_name_dict , product_info_dict , product_name_dict):
    today = datetime.now().strftime("%m_%d_%Y")
    yesterday = datetime.strftime(datetime.now() - timedelta(1), "%m_%d_%Y")
    yesterday_file_error = False
    
    product_change = 0
    product_remove = 0
    product_add = 0
    spec_add = 0
    spec_remove = 0
    
    try:
        old_product_dict = load_obj('./yahoo/' + yesterday + '/product_dict.pkl')
        old_product_supplier_name_dict = load_obj('./yahoo/' + yesterday + '/product_supplier_name_dict.pkl')
        old_product_info_dict = load_obj('./yahoo/' + yesterday + '/product_info_dict.pkl')
        old_product_name_dict = load_obj('./yahoo/' + yesterday + '/product_name_dict.pkl')
    except:
        yesterday_file_error = True

    
    gray = openpyxl.styles.colors.Color(rgb='00E6E6E6')
    fill_gray = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=gray)
    green = openpyxl.styles.colors.Color(rgb='0084EB00')
    fill_green = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=green)
#     record row 
    sheet_info = {}
#     record need to copy cell's row
    be_copy_cell_rows = {}

    column_width = {1:70,2:15,3:40,4:10,5:10,6:10,7:50}
    red = Font(color=colors.RED)
    workbook  = Workbook()
    workbook_changed = Workbook()
    
    for product_key , specs in product_dict.items():
        product_id = product_key
        store_name = product_supplier_name_dict[product_key]
        sheet_name = store_name
        product_name = product_name_dict[product_key]
        try:
            sheet_changed = workbook_changed.get_sheet_by_name(sheet_name)
            sheet = workbook.get_sheet_by_name(sheet_name)
        except:
            sheet = workbook.create_sheet(sheet_name)
            sheet = workbook.get_sheet_by_name(sheet_name)
            
            sheet_changed = workbook_changed.create_sheet(sheet_name)
            sheet_changed = workbook_changed.get_sheet_by_name(sheet_name)
            
            sheet['A1'].value = '商品名稱'
            sheet['B1'].value = '商品編號'
            sheet['C1'].value = '規格'
            sheet['D1'].value = '標價'
            sheet['E1'].value = '售價'
            sheet['F1'].value = '數量'
            sheet['G1'].value = 'info'
            sheet_info[sheet_name] = 2
            
#             for 另外excel
            be_copy_cell_rows[sheet_name] = [1]
            
        need_to_copy = False
    
        row = sheet_info[sheet_name]
        col = 1
        


        sheet.cell(row = row,column = col).value = product_name
        product_info = product_info_dict[product_key]
        sheet.cell(row = row,column = 7).value = product_info

        
        col += 1
        sheet.cell(row = row,column = col).value = product_id
        col += 1
        row_start  = row
        original_spec_add = spec_add
        
        for spec_name , spec in specs.items():
            sheet.cell(row = row,column = col).value = spec_name
            col += 1
            sheet.cell(row = row,column = col).value = spec["original_price"]
            col += 1
            try:
                old_price = int(old_product_dict[product_key][spec_name]['discount_price'])
                price = int(spec["discount_price"])
                if(old_price != price):
                    product_change += 1
                    need_to_copy = True
                    be_copy_cell_rows[sheet_name].append(row)
                    sheet.cell(row = row,column = col).value = str(old_price) + "->" +  str(price)
                    sheet['E'+str(row)].font = red
                else:
                    sheet.cell(row = row, column = col).value = price
                    
            except:
                try:
                    if (old_product_dict[product_key][spec_name]['discount_price'] == '' or  old_product_dict[product_key][spec_name]['discount_price']==None) and spec["discount_price"] != '':
                       # spec_add += 1
                        need_to_copy = True
                        be_copy_cell_rows[sheet_name].append(row)
                        for i in range(1,8):
                            sheet.cell(row=row, column=i).fill = fill_green
                except:
                   # spec_add += 1
                    need_to_copy = True
                    be_copy_cell_rows[sheet_name].append(row)
                    for i in range(1,8):
                        sheet.cell(row=row, column=i).fill = fill_green

                sheet.cell(row=row, column=col).value = spec["discount_price"]
                

            col += 1
            sheet.cell(row = row,column = col).value = spec["quantity"]
            
#             品項新增
            try:
                old_spec = old_product_dict[product_key][spec_name]
            except:
                spec_add += 1
                need_to_copy = True
                be_copy_cell_rows[sheet_name].append(row)
                for i in range(1,8):
                    sheet.cell(row = row,column = i).fill = fill_green
            row += 1
            col = 3

        if(original_spec_add != spec_add):
            product_add += 1
        
        if need_to_copy :
            if row_start not in be_copy_cell_rows[sheet_name]:
                be_copy_cell_rows[sheet_name].append(row_start)
            
      
        sheet_info[sheet_name] = row
        for r in range(1,row):
            for c in range(1,8):
                sheet.cell(row = r,column = c).font = sheet.cell(row = r,column = c).font.copy(size=14)
                sheet.column_dimensions[get_column_letter(c)].width = column_width[c]       
        
        
    
    
    if not yesterday_file_error:
        for product_key , specs in old_product_dict.items():
    #         昨天有，今天沒有，已下架
            try:
                spec_remove_flag = False
                buf = product_dict[product_key]
                spec_name_list = []
                spec_list = []
                for spec_name , spec in specs.items():
                    try:
                        buf = product_dict[product_key][spec_name]

                    except:
                        spec_name_list.append(spec_name)
                        spec_list.append(spec)
                        spec_remove += 1
                        continue
                    if(len(spec_list)!=0):
                        spec_remove_flag = True
                        raise
                        
                        
                
            except:
                product_id = product_key
                store_name = old_product_supplier_name_dict[product_key]
                sheet_name = store_name
                product_name = old_product_name_dict[product_key]
                product_remove += 1

                try:
                    sheet = workbook.get_sheet_by_name(sheet_name)
                except:
                    sheet = workbook.create_sheet(sheet_name)
                    sheet = workbook.get_sheet_by_name(sheet_name)
                    
                    sheet_changed = workbook_changed.create_sheet(sheet_name)
                    sheet_changed = workbook_changed.get_sheet_by_name(sheet_name)
                    
                    sheet['A1'].value = '商品名稱'
                    sheet['B1'].value = '商品編號'
                    sheet['C1'].value = '規格'
                    sheet['D1'].value = '標價'
                    sheet['E1'].value = '售價'
                    sheet['F1'].value = '數量'
                    sheet['G1'].value = 'info'
                    sheet_info[sheet_name] = 2
                    
                    be_copy_cell_rows[sheet_name] = [1]


                row = sheet_info[sheet_name]
                row_start = row
                col = 1
                sheet.cell(row = row,column = col).value = product_name
                sheet.cell(row = row,column = 7).value = old_product_info_dict[product_key]
                col += 1
                sheet.cell(row = row,column = col).value = product_id
                col += 1
                
                
                
                
                
                if(spec_remove_flag):
                    
                    
                    for i in range(len(spec_name_list)):
                        spec = spec_list[i]
                        spec_name = spec_name_list[i]
                        sheet.cell(row = row,column = col).value = spec_name
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["original_price"]
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["discount_price"]
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["quantity"]
                        row += 1
                        col = 3
                else:


                
                    for spec_name , spec in specs.items():
                        sheet.cell(row = row,column = col).value = spec_name
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["original_price"]
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["discount_price"]
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["quantity"]
                        row += 1
                        col = 3
                        spec_remove += 1
                sheet_info[sheet_name] = row

                for r in range(row_start,row):
                    for c in range(1,8):
                        sheet.cell(row = r,column = c).font = sheet.cell(row = r,column = c).font.copy(size=14)
                        sheet.cell(row = r,column = c).fill = fill_gray
                        sheet.column_dimensions[get_column_letter(c)].width = column_width[c]

                for i in range(row_start,row):
                    be_copy_cell_rows[sheet_name].append(i)


                
    try:
        del workbook['Sheet']
        del workbook_changed['Sheet']
    except:
        pass
    
    
    for s_name, rows in be_copy_cell_rows.items():
        sheet_changed = workbook_changed.get_sheet_by_name(s_name)
        sheet = workbook.get_sheet_by_name(s_name)
        rows = list(set(rows))
        i = 1
        for row in rows:
            for col in range(1,8):
                sheet_changed.cell(row = i,column = col).value = sheet.cell(row = row,column = col).value
                sheet_changed.cell(row = i,column = col).font = sheet.cell(row = row,column = col).font.copy()
                sheet_changed.cell(row = i,column = col).fill = sheet.cell(row = row,column = col).fill.copy()
                sheet_changed.column_dimensions[get_column_letter(col)].width = column_width[col]

            i += 1
            
    
    
    
    workbook.save('./yahoo/'+ today + '/yahoo.xlsx')
    workbook_changed.save('./yahoo/'+ today + '/yahoo_changed.xlsx')
    fileNames = ['yahoo/'+ today + '/yahoo.xlsx',
                 'yahoo/'+ today + '/yahoo_changed.xlsx']
    
    today = datetime.now().strftime("%m_%d_%Y %H:%M:%S")
    text = today + ' 商品變更數量:' + str(product_change) + ', 商品移除數量: ' + (str(product_remove))
    text += ' ,品項移除數量:' + str(spec_remove)
    text += '\n總共有' + str(product_add) + '個商品的品項有新增，共比昨天多了'+ str(spec_add) + '個品項'
    text += '\n綠色代表新增，灰色代表移除，紅色代表變更！'
    
    return fileNames, text
    
def send_mail(send_from, send_to, subject, text, files=None,server="127.0.0.1"):
    assert isinstance(send_to, list)
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = COMMASPACE.join(send_to)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    msg.attach(MIMEText(text))
    
    if files is not None:
        for file in files:
            with open(file, "rb") as fil:
                part = MIMEApplication(
                    fil.read(),
                    Name=basename(file)
                )
                # After the file is closed
                part['Content-Disposition'] = 'attachment; filename="%s"' % basename(file)
                msg.attach(part)
    smtp = smtplib.SMTP('10.210.1.221')
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.close()




subject = 'Yahoo商品爬蟲'
send_to = ['ruby.lin@e7line.com',
           'xing.chen@gigabyte.com',
           'chaoyang.huang@gigabyte.com',
	   'harrychiang0@gmail.com',
	   'carina.wang@e7line.com',
	   'kelsey.chang@e7line.com',
	   'vivian.hung@e7line.com']

#send_to = ['harrychiang0@gmail.com']

test_api = 'https://www.e7line.com:8080/spiderdata3.aspx'
api = 'https://www.e7line.com/spiderdata3.aspx'
with urllib.request.urlopen(api) as url:
    datas = json.loads(url.read().decode())
    
try:
    yahoo_data  = datas[0]
except:
    text = '今天yahoo商品沒有資訊！'
    send_mail('e7line@gigabyte.com', send_to , subject, text , files=None,server="127.0.0.1")

    print('cant get yahoo data')
    raise SystemExit("stop program")

product_dict, product_supplier_name_dict , product_info_dict , product_name_dict = processYahooData(yahoo_data)
fileNames , text = dumpExcel(product_dict, product_supplier_name_dict , product_info_dict , product_name_dict)

    
# send_to = ['harrychiang0@gmail.com']
send_mail('e7line@gigabyte.com', send_to , subject, text , files=fileNames,server="127.0.0.1")






