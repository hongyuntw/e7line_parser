from bs4 import BeautifulSoup
import os
import re
from datetime import datetime, timedelta
import requests
import json
import urllib.request, json 
import ast 
import demjson
from fake_useragent import UserAgent
from selenium import webdriver
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import openpyxl

from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
import selenium.webdriver.support.ui as ui
from time import sleep
from openpyxl.styles import Alignment


import pickle
import errno
from openpyxl.utils import get_column_letter

import smtplib
from os.path import basename
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate
from collections import OrderedDict
def load_obj(path):
    with open(path, 'rb') as f:
        return pickle.load(f)


def dumpExcel(product_dict, product_supplier_name_dict , product_info_dict , product_name_dict):
    today = datetime.now().strftime("%m_%d_%Y")
    yesterday = datetime.strftime(datetime.now() - timedelta(1), "%m_%d_%Y")
    yesterday_file_error = False
    
    product_change = 0
    product_remove = 0
    product_add = 0
    spec_add = 0
    spec_remove = 0
    
    try:
        old_product_dict = load_obj('./yahoo/' + yesterday + '/product_dict.pkl')
        old_product_supplier_name_dict = load_obj('./yahoo/' + yesterday + '/product_supplier_name_dict.pkl')
        old_product_info_dict = load_obj('./yahoo/' + yesterday + '/product_info_dict.pkl')
        old_product_name_dict = load_obj('./yahoo/' + yesterday + '/product_name_dict.pkl')
    except:
        yesterday_file_error = True

    
    gray = openpyxl.styles.colors.Color(rgb='00E6E6E6')
    fill_gray = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=gray)
    green = openpyxl.styles.colors.Color(rgb='0084EB00')
    fill_green = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=green)
     #record row 
    sheet_info = {}
     #record need to copy cell's row
    be_copy_cell_rows = {}

    column_width = {1:70,2:15,3:40,4:10,5:10,6:10,7:50}
    red = Font(color=colors.RED)
    workbook  = Workbook()
    workbook_changed = Workbook()
    
    for product_key, specs in product_dict.items():
        # if product_key == '100738571990':
        #     print(123)
        product_id = product_key
        store_name = product_supplier_name_dict[product_key]
        sheet_name = store_name
        product_name = product_name_dict[product_key]
        try:
            sheet_changed = workbook_changed.get_sheet_by_name(sheet_name)
            sheet = workbook.get_sheet_by_name(sheet_name)
        except:
            sheet = workbook.create_sheet(sheet_name)
            sheet = workbook.get_sheet_by_name(sheet_name)
            
            sheet_changed = workbook_changed.create_sheet(sheet_name)
            sheet_changed = workbook_changed.get_sheet_by_name(sheet_name)
            
            sheet['A1'].value = '商品名稱'
            sheet['B1'].value = '商品編號'
            sheet['C1'].value = '規格'
            sheet['D1'].value = '標價'
            sheet['E1'].value = '售價'
            sheet['F1'].value = '數量'
            sheet['G1'].value = 'info'
            sheet_info[sheet_name] = 2
            
            #for 另外excel
            be_copy_cell_rows[sheet_name] = [1]
            
        need_to_copy = False
    
        row = sheet_info[sheet_name]
        col = 1
        


        sheet.cell(row = row,column = col).value = product_name
        product_info = product_info_dict[product_key]
        sheet.cell(row = row,column = 7).value = product_info

        
        col += 1
        sheet.cell(row = row,column = col).value = product_id
        col += 1
        row_start  = row
        original_spec_add = spec_add
        
        for spec_name , spec in specs.items():
            sheet.cell(row = row,column = col).value = spec_name
            col += 1
            sheet.cell(row = row,column = col).value = spec["original_price"]
            col += 1
            try:
                old_price = int(old_product_dict[product_key][spec_name]['discount_price'])
                price = int(spec["discount_price"])
                if(old_price != price):
                    product_change += 1
                    need_to_copy = True
                    be_copy_cell_rows[sheet_name].append(row)
                    sheet.cell(row = row,column = col).value = str(old_price) + "->" +  str(price)
                    sheet['E'+str(row)].font = red
                else:
                    sheet.cell(row = row, column = col).value = price
                    
            except:
                try:
                    if (old_product_dict[product_key][spec_name]['discount_price'] == '' or  old_product_dict[product_key][spec_name]['discount_price']==None) and spec["discount_price"] != '':
                       # spec_add += 1
                        need_to_copy = True
                        be_copy_cell_rows[sheet_name].append(row)
                        for i in range(1,8):
                            sheet.cell(row=row, column=i).fill = fill_green
                    elif spec["discount_price"] == '' and old_product_dict[product_key][spec_name]['discount_price'] != '':
                        spec_remove += 1
                        need_to_copy = True
                        be_copy_cell_rows[sheet_name].append(row)
                        for i in range(1,8):
                            sheet.cell(row=row, column=i).fill = fill_gray

                except:
                   # spec_add += 1
                    need_to_copy = True
                    be_copy_cell_rows[sheet_name].append(row)
                    for i in range(1,8):
                        sheet.cell(row=row, column=i).fill = fill_green

                sheet.cell(row=row, column=col).value = spec["discount_price"]
                

            col += 1
            sheet.cell(row = row,column = col).value = spec["quantity"]
            
             #品項新增
            try:
                old_spec = old_product_dict[product_key][spec_name]
            except:
                spec_add += 1
                need_to_copy = True
                be_copy_cell_rows[sheet_name].append(row)
                for i in range(1,8):
                    sheet.cell(row = row,column = i).fill = fill_green
            row += 1
            col = 3

        if(original_spec_add != spec_add):
            product_add += 1
        
        if need_to_copy :
            if row_start not in be_copy_cell_rows[sheet_name]:
                be_copy_cell_rows[sheet_name].append(row_start)
            
      
        sheet_info[sheet_name] = row
        for r in range(1,row):
            for c in range(1,8):
                sheet.cell(row = r,column = c).font = sheet.cell(row = r,column = c).font.copy(size=14)
                sheet.column_dimensions[get_column_letter(c)].width = column_width[c]       
        
        
    
    
    if not yesterday_file_error:
        for product_key , specs in old_product_dict.items():
             #昨天有，今天沒有，已下架
            try:
                spec_remove_flag = False
                buf = product_dict[product_key]
                spec_name_list = []
                spec_list = []
                for spec_name , spec in specs.items():
                    try:
                        buf = product_dict[product_key][spec_name]

                    except:
                        spec_name_list.append(spec_name)
                        spec_list.append(spec)
                        spec_remove += 1
                        continue
                    if(len(spec_list)!=0):
                        spec_remove_flag = True
                        raise
                        
                        
                
            except:
                product_id = product_key
                store_name = old_product_supplier_name_dict[product_key]
                sheet_name = store_name
                product_name = old_product_name_dict[product_key]
                product_remove += 1

                try:
                    sheet = workbook.get_sheet_by_name(sheet_name)
                except:
                    sheet = workbook.create_sheet(sheet_name)
                    sheet = workbook.get_sheet_by_name(sheet_name)
                    
                    sheet_changed = workbook_changed.create_sheet(sheet_name)
                    sheet_changed = workbook_changed.get_sheet_by_name(sheet_name)
                    
                    sheet['A1'].value = '商品名稱'
                    sheet['B1'].value = '商品編號'
                    sheet['C1'].value = '規格'
                    sheet['D1'].value = '標價'
                    sheet['E1'].value = '售價'
                    sheet['F1'].value = '數量'
                    sheet['G1'].value = 'info'
                    sheet_info[sheet_name] = 2
                    
                    be_copy_cell_rows[sheet_name] = [1]


                row = sheet_info[sheet_name]
                row_start = row
                col = 1
                sheet.cell(row = row,column = col).value = product_name
                sheet.cell(row = row,column = 7).value = old_product_info_dict[product_key]
                col += 1
                sheet.cell(row = row,column = col).value = product_id
                col += 1
                
                
                
                
                
                if(spec_remove_flag):
                    
                    
                    for i in range(len(spec_name_list)):
                        spec = spec_list[i]
                        spec_name = spec_name_list[i]
                        sheet.cell(row = row,column = col).value = spec_name
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["original_price"]
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["discount_price"]
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["quantity"]
                        row += 1
                        col = 3
                else:

                    for spec_name , spec in specs.items():
                        sheet.cell(row = row,column = col).value = spec_name
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["original_price"]
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["discount_price"]
                        col += 1
                        sheet.cell(row = row,column = col).value = spec["quantity"]
                        row += 1
                        col = 3
                        spec_remove += 1
                sheet_info[sheet_name] = row

                for r in range(row_start,row):
                    for c in range(1,8):
                        sheet.cell(row = r,column = c).font = sheet.cell(row = r,column = c).font.copy(size=14)
                        sheet.cell(row = r,column = c).fill = fill_gray
                        sheet.column_dimensions[get_column_letter(c)].width = column_width[c]

                for i in range(row_start,row):
                    be_copy_cell_rows[sheet_name].append(i)


                
    try:
        del workbook['Sheet']
        del workbook_changed['Sheet']
    except:
        pass
    
    
    for s_name, rows in be_copy_cell_rows.items():
        sheet_changed = workbook_changed.get_sheet_by_name(s_name)
        sheet = workbook.get_sheet_by_name(s_name)
        rows = list(set(rows))
        rows = sorted(rows)
        i = 1
        for row in rows:
            for col in range(1,8):
                sheet_changed.cell(row = i,column = col).value = sheet.cell(row = row,column = col).value
                sheet_changed.cell(row = i,column = col).font = sheet.cell(row = row,column = col).font.copy()
                sheet_changed.cell(row = i,column = col).fill = sheet.cell(row = row,column = col).fill.copy()
                sheet_changed.column_dimensions[get_column_letter(col)].width = column_width[col]

            i += 1
            
    
    
    
    workbook.save('./yahoo/'+ today + '/yahoo.xlsx')
    workbook_changed.save('./yahoo/'+ today + '/yahoo_changed.xlsx')
    fileNames = ['yahoo/'+ today + '/yahoo.xlsx',
                 'yahoo/'+ today + '/yahoo_changed.xlsx']
    
    today = datetime.now().strftime("%m_%d_%Y %H:%M:%S")
    text = today + ' 商品變更數量:' + str(product_change) + ', 商品移除數量: ' + (str(product_remove))
    text += ' ,品項移除數量:' + str(spec_remove)
    text += '\n總共有' + str(product_add) + '個商品的品項有新增，共比昨天多了'+ str(spec_add) + '個品項'
    text += '\n綠色代表新增，灰色代表移除，紅色代表變更！'
    
    return fileNames, text


today = datetime.now().strftime("%m_%d_%Y")
product_dict = load_obj('./yahoo/' + today + '/product_dict.pkl')
product_supplier_name_dict = load_obj('./yahoo/' + today + '/product_supplier_name_dict.pkl')
product_info_dict = load_obj('./yahoo/' + today + '/product_info_dict.pkl')
product_name_dict = load_obj('./yahoo/' + today + '/product_name_dict.pkl')
dumpExcel(product_dict, product_supplier_name_dict , product_info_dict , product_name_dict)
    